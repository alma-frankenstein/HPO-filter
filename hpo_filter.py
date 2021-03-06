#!/usr/bin/env python3
# -*- coding: utf-8 -*-


import openpyxl

#make sure you're in the directory where the HPO spreadsheet is saved
workbook_name = input("Please enter name of spreadsheet, i.e. 'n17-00138 HPO List.xlsx': ")
#include file extension (example: .xlsx). Don't put in quotes.

wb = openpyxl.load_workbook(workbook_name)

sheet_name = input("Please enter name of sheet that contains the HPO terms and genes, i.e. 'HPO Genes': ")
sheet = wb[sheet_name]
#this needs to match the name of the Excel spreadsheet

y = sheet.max_row
row_range_cutoff = y+1

x = sheet.max_column
col_range_cutoff = x + 1

starting_column = 1
HPO_terms = []

genes = input("Please input genes: ")
genes = genes.split()


while starting_column < col_range_cutoff:
    column = []
#for each row in the first column, take the value and put it into an empty list.
#if an imput gene is present, put the HPO term into the HPO_terms list.
#do this for every column
    for roww in range(1,row_range_cutoff):
        z=(sheet.cell(row = roww, column=starting_column).value)
        column.append(z)
    starting_column += 1
    for gene in genes:
        if gene in column:
            HPO_terms.append(column[1])
            #if HPO term isn't in row 2 of spreadsheet, change the index to match


#remove duplicates by changing to a dictionary then back
HPO_terms = list(dict.fromkeys(HPO_terms))

print(HPO_terms)
